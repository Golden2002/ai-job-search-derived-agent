# -*- coding: utf-8 -*-
"""resume_product.render.xlsx_import —— Excel 简历 → openpyxl 单元格样式 → HTML 表格。

Oracle 论证 P1：Excel 是表格结构化，openpyxl 能读取单元格的完整样式
（字体/字号/加粗/颜色/填充/对齐/边框）+ merged_cells，可精准还原为 HTML 表格。

依赖：openpyxl（可选，无则优雅降级返回空）。
"""
from __future__ import annotations

from typing import Any, Dict, List, Tuple


def _esc(s) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _rgb_hex(color) -> str:
    """openpyxl Color 对象 → '#RRGGBB'（仅 rgb 类型；theme/indexed 跳过）。"""
    try:
        if color is None:
            return ""
        if getattr(color, "type", None) != "rgb":
            return ""
        rgb = getattr(color, "rgb", None)
        if not rgb or not isinstance(rgb, str) or len(rgb) < 6:
            return ""
        return "#" + rgb[-6:]
    except Exception:
        return ""


def _parse_merges(ws) -> Tuple[Dict[Tuple[int, int], Tuple[int, int]], set]:
    """合并单元格 → {起始坐标: (rowspan, colspan)} + 被覆盖坐标集合。"""
    merges: Dict[Tuple[int, int], Tuple[int, int]] = {}
    skip: set = set()
    for rng in ws.merged_cells.ranges:
        key = (rng.min_row, rng.min_col)
        merges[key] = (rng.max_row - rng.min_row + 1,
                       rng.max_col - rng.min_col + 1)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != key:
                    skip.add((r, c))
    return merges, skip


def import_resume_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """Excel 简历 → {html, css, meta}（HTML 表格还原单元格样式）。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"html": "", "css": "", "meta": {"paragraphs": 0, "css_length": 0}}

    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
    except Exception:
        return {"html": "", "css": "", "meta": {"paragraphs": 0, "css_length": 0}}

    merges, skip = _parse_merges(ws)

    rows_html: List[str] = []
    for r in range(1, ws.max_row + 1):
        cells_html = ""
        for c in range(1, ws.max_column + 1):
            if (r, c) in skip:
                continue
            cell = ws.cell(r, c)
            value = cell.value
            if value is None:
                value = ""
            style = []
            font = cell.font
            if font.name:
                style.append(f"font-family:'{font.name}'")
            if font.size:
                style.append(f"font-size:{font.size}pt")
            if font.bold:
                style.append("font-weight:bold")
            if font.italic:
                style.append("font-style:italic")
            fg = _rgb_hex(font.color)
            if fg:
                style.append(f"color:{fg}")
            # 填充色（背景）——仅 solid 填充才还原，默认无填充跳过
            try:
                fill = cell.fill
                if fill is not None and getattr(fill, "patternType", None) == "solid":
                    bg = _rgb_hex(fill.fgColor)
                    if bg:
                        style.append(f"background:{bg}")
            except Exception:
                pass
            # 对齐
            try:
                ha = cell.alignment.horizontal if cell.alignment else None
                if ha:
                    style.append(f"text-align:{ha}")
            except Exception:
                pass
            # 合并单元格 colspan/rowspan
            attrs = ""
            if (r, c) in merges:
                rs, cs = merges[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            style_str = ";".join(style)
            cells_html += (f'<td style="border:1px solid #999;padding:4pt 6pt;'
                           f'{style_str}"{attrs}>{_esc(str(value))}</td>')
        rows_html.append(f"<tr>{cells_html}</tr>")

    html = ('<table style="border-collapse:collapse;width:100%;margin:6pt 0;">'
            + "".join(rows_html) + "</table>")
    css = ("@page { size: A4; margin: 18mm 16mm; }\n"
           "body { font-family: 'SimSun','Songti SC','Noto Serif CJK SC',serif; }")

    meta = {
        "paragraphs": ws.max_row,
        "columns": ws.max_column,
        "merged_cells": len(merges),
        "css_length": len(css),
        "xlsx_engine": "openpyxl",
    }
    return {"html": html, "css": css, "meta": meta, "parsed": {"sheets": wb.sheetnames}}
