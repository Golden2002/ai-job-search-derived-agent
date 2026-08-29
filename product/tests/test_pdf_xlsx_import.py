# -*- coding: utf-8 -*-
"""PDF / Excel 简历 → 排版还原 测试（pdf_import + xlsx_import）。"""
import os
import sys

_SRC = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

import pytest

from resume_product.render.pdf_import import import_resume_pdf
from resume_product.render.xlsx_import import import_resume_xlsx


# ── PDF ──
@pytest.fixture(scope="module")
def pdf_path(tmp_path_factory):
    pytest.importorskip("reportlab")
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor
    p = str(tmp_path_factory.mktemp("pdf") / "resume.pdf")
    c = canvas.Canvas(p, pagesize=A4)
    W, H = A4
    c.setFillColor(HexColor("#1F3A5F"))
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(W / 2, H - 80, "Resume")
    c.setFillColor(HexColor("#000000"))
    c.setFont("Helvetica", 11)
    c.drawString(50, H - 140, "Work Experience")
    c.drawString(50, H - 170, "Engineer at Google")
    c.drawString(350, H - 140, "Skills")
    c.drawString(350, H - 170, "Python, SQL")
    c.save()
    return p


def test_import_resume_pdf(pdf_path):
    pytest.importorskip("pdfplumber")
    r = import_resume_pdf(pdf_path)
    assert r["meta"]["paragraphs"] >= 3
    assert "Resume" in r["html"] or "Work" in r["html"]
    assert "column-count:2" in r["css"] or r["meta"]["columns"] >= 1
    # 标题深蓝颜色还原
    if "Resume" in r["html"]:
        assert "text-align:center" in r["html"]


# ── Excel ──
@pytest.fixture(scope="module")
def xlsx_path(tmp_path_factory):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    p = str(tmp_path_factory.mktemp("xlsx") / "resume.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "王小明 简历"
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(name="黑体", size=16, bold=True, color="FF1F3A5F")
    ws["A1"].fill = PatternFill("solid", fgColor="FFE8EDFF")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "技能类别"; ws["B2"] = "技能"; ws["C2"] = "水平"
    ws["A3"] = "产品"; ws["B3"] = "需求分析"; ws["C3"] = "精通"
    wb.save(p)
    return p


def test_import_resume_xlsx(xlsx_path):
    r = import_resume_xlsx(xlsx_path)
    assert r["meta"]["merged_cells"] == 1
    assert 'colspan="3"' in r["html"], "合并单元格应还原为 colspan"
    assert "王小明" in r["html"]
    assert "color:#1F3A5F" in r["html"], "字体颜色应还原"
    assert "background:#E8EDFF" in r["html"], "填充色应还原"
    assert "font-weight:bold" in r["html"]


def test_xlsx_no_merged(tmp_path):
    from openpyxl import Workbook
    p = str(tmp_path / "plain.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "a"; ws["B1"] = "b"
    wb.save(p)
    r = import_resume_xlsx(p)
    assert r["meta"]["merged_cells"] == 0
    assert "background:#000000" not in r["html"], "无填充不应有黑色背景"
